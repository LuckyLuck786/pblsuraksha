"""
Management command: import_test_cases
======================================
Imports 300 research cases from the project Excel file into the Complaints DB.

Expected Excel columns (row 1 = headers):
  title | description | category | priority | severity_score | incident_location | status

Usage:
  python manage.py import_test_cases
  python manage.py import_test_cases --file /path/to/cases.xlsx
  python manage.py import_test_cases --clear-all    # wipe ALL complaints first
  python manage.py import_test_cases --clear        # wipe only TEST_IMPORT rows first
"""

import os
import random
from django.core.management.base import BaseCommand, CommandError
from django.contrib.auth import get_user_model
import openpyxl


VALID_CATEGORIES = [
    'theft', 'assault', 'harassment', 'traffic', 'fraud',
    'cybercrime', 'domestic', 'missing_person', 'drug_activity',
    'vandalism', 'noise', 'other',
]
VALID_PRIORITIES = ['low', 'medium', 'high', 'critical']
VALID_STATUSES   = ['pending', 'acknowledged', 'in_progress', 'resolved', 'closed', 'rejected']

# ── Bangalore area geocoordinate catalogue ────────────────────────────────────
_AREAS = [
    ('ulsoor',              12.9811, 77.6206),
    ('hennur',              13.0271, 77.6400),
    ('malleshwaram',        13.0030, 77.5688),
    ('koramangala',         12.9352, 77.6245),
    ('indiranagar',         12.9784, 77.6408),
    ('whitefield',          12.9698, 77.7500),
    ('electronic city',     12.8459, 77.6622),
    ('jayanagar',           12.9308, 77.5838),
    ('jp nagar',            12.9081, 77.5888),
    ('hsr layout',          12.9116, 77.6473),
    ('btm layout',          12.9166, 77.6101),
    ('marathahalli',        12.9588, 77.6973),
    ('hebbal',              13.0348, 77.5942),
    ('yelahanka',           13.0998, 77.5964),
    ('rajajinagar',         12.9980, 77.5489),
    ('basavanagudi',        12.9427, 77.5740),
    ('banashankari',        12.9260, 77.5561),
    ('rt nagar',            13.0219, 77.5961),
    ('frazer town',         12.9830, 77.6202),
    ('shivajinagar',        12.9860, 77.5970),
    ('mg road',             12.9752, 77.6057),
    ('brigade road',        12.9716, 77.6066),
    ('majestic',            12.9762, 77.5713),
    ('domlur',              12.9607, 77.6382),
    ('bellandur',           12.9257, 77.6767),
    ('sarjapur',            12.9060, 77.6855),
    ('cv raman nagar',      12.9943, 77.6607),
    ('kr puram',            13.0009, 77.6933),
    ('banaswadi',           13.0103, 77.6536),
    ('sadashivanagar',      13.0063, 77.5750),
    ('yeshwanthpur',        13.0234, 77.5498),
    ('peenya',              13.0285, 77.5208),
    ('nagarbhavi',          12.9517, 77.5138),
    ('vijayanagar',         12.9744, 77.5358),
    ('kengeri',             12.9074, 77.4853),
    ('mysore road',         12.9618, 77.5245),
    ('rajarajeshwari',      12.9233, 77.5059),
    ('brookefield',         12.9725, 77.7010),
    ('kundalahalli',        12.9730, 77.7121),
    ('richmond town',       12.9622, 77.5979),
    ('langford town',       12.9570, 77.5993),
    ('vasanth nagar',       12.9932, 77.5876),
    ('seshadripuram',       12.9943, 77.5699),
    ('jayanagar 4th block', 12.9245, 77.5855),
    ('rajiv gandhi nagar',  12.9690, 77.5425),
    ('uttarahalli',         12.8968, 77.5424),
    ('gottigere',           12.8730, 77.6024),
    ('begur',               12.8780, 77.6299),
    ('hulimavu',            12.8891, 77.6205),
    ('silk board',          12.9176, 77.6223),
]

_CATEGORY_AREA_BIAS = {
    'theft'         : ['mg road', 'koramangala', 'brigade road', 'marathahalli', 'whitefield'],
    'assault'       : ['koramangala', 'indiranagar', 'ulsoor', 'mg road', 'frazer town'],
    'harassment'    : ['mg road', 'majestic', 'indiranagar', 'koramangala', 'brigade road'],
    'traffic'       : ['whitefield', 'marathahalli', 'electronic city', 'kr puram', 'hebbal'],
    'fraud'         : ['whitefield', 'marathahalli', 'koramangala', 'indiranagar', 'hsr layout'],
    'cybercrime'    : ['whitefield', 'electronic city', 'koramangala', 'hsr layout', 'btm layout'],
    'domestic'      : ['jayanagar', 'basavanagudi', 'banashankari', 'jp nagar', 'rajajinagar'],
    'missing_person': ['majestic', 'kr puram', 'banaswadi', 'hebbal', 'yelahanka'],
    'drug_activity' : ['majestic', 'frazer town', 'shivajinagar', 'ulsoor', 'banaswadi'],
    'vandalism'     : ['jayanagar', 'btm layout', 'hsr layout', 'jp nagar', 'banashankari'],
    'noise'         : ['koramangala', 'indiranagar', 'jp nagar', 'btm layout', 'hsr layout'],
    'other'         : ['mg road', 'jayanagar', 'shivajinagar', 'malleshwaram', 'hebbal'],
}


def _pick_coords(location_str: str, category: str) -> tuple:
    loc = (location_str or '').lower()
    for frag, lat, lon in _AREAS:
        if frag in loc:
            return (
                round(lat + random.uniform(-0.006, 0.006), 6),
                round(lon + random.uniform(-0.006, 0.006), 6),
            )
    bias = _CATEGORY_AREA_BIAS.get(category, ['mg road'])
    chosen = random.choice(bias)
    for frag, lat, lon in _AREAS:
        if frag == chosen:
            return (
                round(lat + random.uniform(-0.012, 0.012), 6),
                round(lon + random.uniform(-0.012, 0.012), 6),
            )
    return (
        round(12.9716 + random.uniform(-0.04, 0.04), 6),
        round(77.5946 + random.uniform(-0.04, 0.04), 6),
    )


class Command(BaseCommand):
    help = 'Import 300 research cases from the project Excel file.'

    def add_arguments(self, parser):
        parser.add_argument(
            '--file',
            default=os.path.expanduser('~/Downloads/actual 300 cases in detail.xlsx'),
            help='Path to the Excel file.',
        )
        parser.add_argument(
            '--clear-all', action='store_true',
            help='Delete ALL complaints from the database before importing.',
        )
        parser.add_argument(
            '--clear', action='store_true',
            help='Delete only TEST_IMPORT-tagged complaints before importing.',
        )

    def handle(self, *args, **options):
        from apps.complaints.models import Complaint
        from apps.intelligence.engine import compute_severity

        User      = get_user_model()
        file_path = options['file']
        random.seed(42)

        if not os.path.exists(file_path):
            raise CommandError(f'File not found: {file_path}')

        # ── Clear ─────────────────────────────────────────────────────────────
        if options['clear_all']:
            deleted, _ = Complaint.objects.all().delete()
            self.stdout.write(self.style.WARNING(f'Deleted ALL {deleted} existing complaints.'))
        elif options['clear']:
            deleted, _ = Complaint.objects.filter(
                ai_summary__startswith='TEST_IMPORT|'
            ).delete()
            self.stdout.write(self.style.WARNING(f'Cleared {deleted} TEST_IMPORT complaints.'))

        # ── Citizen users for round-robin assignment ───────────────────────────
        citizens = list(User.objects.filter(role='citizen').order_by('id'))
        if not citizens:
            citizens = list(User.objects.filter(is_superuser=False)[:3])
        if not citizens:
            citizens = [User.objects.filter(is_superuser=True).first()]
        self.stdout.write(
            f'Assigning to {len(citizens)} citizen(s): '
            + ', '.join(u.username for u in citizens)
        )

        # ── Status spread (realistic lifecycle) ───────────────────────────────
        status_pool = (
            ['pending'] * 4 +
            ['acknowledged'] * 2 +
            ['in_progress'] * 2 +
            ['resolved'] * 1 +
            ['closed'] * 1
        ) * 30   # 300 slots
        random.shuffle(status_pool)

        # ── Load Excel ────────────────────────────────────────────────────────
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active

        # Read header row to find column positions
        headers = [str(ws.cell(1, c).value or '').strip().lower()
                   for c in range(1, ws.max_column + 1)]
        self.stdout.write(f'Excel columns: {headers}')

        def col(name):
            try:
                return headers.index(name)
            except ValueError:
                return None

        idx_title    = col('title')
        idx_desc     = col('description')
        idx_cat      = col('category')
        idx_pri      = col('priority')
        idx_sev      = col('severity_score')
        idx_loc      = col('incident_location')
        idx_status   = col('status')

        rows = list(ws.iter_rows(min_row=2, values_only=True))
        self.stdout.write(f'Found {len(rows)} data rows.')

        created = skipped = errors = 0

        for i, row in enumerate(rows):
            try:
                def get(idx, default=''):
                    if idx is None or idx >= len(row):
                        return default
                    v = row[idx]
                    return str(v).strip() if v is not None else default

                title       = get(idx_title)
                description = get(idx_desc)
                category    = get(idx_cat, 'other').lower()
                priority    = get(idx_pri, 'medium').lower()
                location    = get(idx_loc, 'Bangalore, Karnataka')
                row_status  = get(idx_status, 'pending').lower()

                if not title or not description:
                    self.stdout.write(self.style.WARNING(f'  Row {i+2}: empty title/description — skipped'))
                    skipped += 1
                    continue

                if category not in VALID_CATEGORIES:
                    category = 'other'
                if priority not in VALID_PRIORITIES:
                    priority = 'medium'
                if row_status not in VALID_STATUSES:
                    row_status = 'pending'

                # Use spreadsheet status if it's not all-pending, else use spread pool
                final_status = row_status if row_status != 'pending' else status_pool[i]

                severity = compute_severity(title, description, category, priority)
                lat, lon = _pick_coords(location, category)
                reporter  = citizens[i % len(citizens)]

                Complaint.objects.create(
                    title             = title[:300],
                    description       = description,
                    category          = category,
                    priority          = priority,
                    severity_score    = severity,
                    incident_location = location or 'Bangalore, Karnataka',
                    incident_address  = location,
                    latitude          = lat,
                    longitude         = lon,
                    reporter          = reporter,
                    status            = final_status,
                    is_anonymous      = False,
                    ai_category       = category,
                    ai_priority       = priority,
                    ai_summary        = (
                        f'TEST_IMPORT|{category}|{priority}|'
                        f'Research dataset case. Category: {category}, Priority: {priority}.'
                    ),
                )
                created += 1
                if created % 50 == 0:
                    self.stdout.write(f'  … {created} imported so far')

            except Exception as exc:
                self.stdout.write(self.style.ERROR(f'  Row {i+2}: ERROR — {exc}'))
                errors += 1

        # ── Summary ───────────────────────────────────────────────────────────
        self.stdout.write(self.style.SUCCESS(
            f'\nDone!  Created={created}  Skipped={skipped}  Errors={errors}\n'
            f'Cases tagged with ai_summary prefix "TEST_IMPORT|" for easy filtering.\n'
            f'All cases geocoded with Bengaluru coordinates — hotspot map is ready.'
        ))

        # Per-user breakdown
        for u in citizens:
            n = Complaint.objects.filter(reporter=u, ai_summary__startswith='TEST_IMPORT|').count()
            self.stdout.write(f'  {u.username}: {n} cases')
